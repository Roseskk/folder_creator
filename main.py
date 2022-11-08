from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
from openpyxl import load_workbook
from pathlib import Path

import traceback
import utils as utl


gauth = GoogleAuth()
gauth.LocalWebserverAuth()


drive = GoogleDrive(gauth)

# Create path choice
path_to_file = input('Path to file: ')
try:
    if Path(path_to_file).exists():
        data = load_workbook(f'{Path(path_to_file)}')

        # Смотрим профиль
        profile = utl.list_folder(drive, '1aNNJp32P5nuVyKq5fG3-2Qmy4Rke-UPJ')

        for name_profile in profile:
            print(name_profile)

        title_handler = input('Profile ---->>>>>')

        # Смотрим курс
        courses = utl.list_course(title_handler, profile, drive)

        # Выводим курсы
        for course in courses:
            print(course)

        course_handler = input("Course ------->>>>")

        groups_list = utl.list_group_with_id(course_handler, courses, drive)

        # Тут ID курса смотрим на гугл диске
        for group in groups_list.items():
            groups_tuple_course_id = group[-1]

        group_handler_values = dict()

        # Название листов в файле
        for excel_list_name in data.sheetnames:
            print('Excel pages(groups): ', excel_list_name)

        group_handler = input("Group ------>>>>")

        while group_handler not in data.sheetnames:
            group_handler = input("Group ------>>>>")

        # Создаем группу в указаном Курсе
        utl.create_group_folder(groups_tuple_course_id, group_handler, drive)

        # Получаем ID гугл диска курса
        updated = utl.list_course(course_handler, courses, drive)

        result_group = str()

        # Берём ID созданной группы
        for group in updated:
            if group_handler in group.values():
                result_group = group.get('id')
                break

        # Создаем папки студентов и правами
        utl.create_student_folder(result_group, group_handler, data, drive)

    else:
        print('Incorrect path')


except Exception as e:
    print('Wrong ID of Google Drive!', traceback.print_exc())



